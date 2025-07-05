import openpyxl
from pathlib import Path

def analyze_excel_file(file_path):
    """Analyze the structure of an Excel file to understand quantity detection issues."""
    
    # Load the workbook
    print(f"Loading file: {file_path}")
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading file: {e}")
        return
    
    # Get the active sheet
    sheet = workbook.active
    print(f"\nActive sheet name: {sheet.title}")
    print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
    
    # Print header row (assuming it's row 1)
    print("\n" + "="*80)
    print("COLUMN HEADERS (Row 1):")
    print("="*80)
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(f"Col {col}: {cell_value}")
    
    for header in headers:
        print(header)
    
    # Analyze first 20 rows
    print("\n" + "="*80)
    print("FIRST 20 ROWS ANALYSIS:")
    print("="*80)
    
    rows_with_quantities = []
    quantity_columns = set()
    
    for row_num in range(1, min(21, sheet.max_row + 1)):
        print(f"\nRow {row_num}:")
        row_values = []
        non_empty_cells = []
        
        # Check each cell in the row
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_num, column=col)
            value = cell.value
            
            if value is not None and str(value).strip() != "":
                row_values.append(f"Col {col}: {value}")
                non_empty_cells.append(col)
                
                # Check if this might be a quantity
                if isinstance(value, (int, float)) and value > 0:
                    # Check if this isn't a price column (usually has decimal places or large values)
                    if not (isinstance(value, float) and value > 100) or value == int(value):
                        quantity_columns.add(col)
        
        # Print row information
        if row_values:
            print(f"  Non-empty cells in columns: {non_empty_cells}")
            for val in row_values:
                print(f"    {val}")
            
            # Check if this row likely has a quantity
            has_quantity = False
            for col in non_empty_cells:
                cell_value = sheet.cell(row=row_num, column=col).value
                if isinstance(cell_value, (int, float)) and cell_value > 0:
                    # Simple heuristic: quantities are usually integers between 1-1000
                    if 0 < cell_value <= 1000 and (isinstance(cell_value, int) or cell_value == int(cell_value)):
                        has_quantity = True
                        rows_with_quantities.append(row_num)
                        break
        else:
            print(f"  [Empty row]")
    
    # Additional analysis for all rows
    print("\n" + "="*80)
    print("FULL FILE QUANTITY ANALYSIS:")
    print("="*80)
    
    all_rows_with_quantities = []
    quantity_patterns = {}
    
    for row_num in range(2, sheet.max_row + 1):  # Skip header row
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row_num, column=col).value
            
            # Look for numeric values that could be quantities
            if isinstance(cell_value, (int, float)) and cell_value > 0:
                # Check if it's likely a quantity (not a price or ID)
                if 0 < cell_value <= 1000 and (isinstance(cell_value, int) or cell_value == int(cell_value)):
                    if row_num not in all_rows_with_quantities:
                        all_rows_with_quantities.append(row_num)
                    
                    # Track which columns contain quantities
                    if col not in quantity_patterns:
                        quantity_patterns[col] = []
                    quantity_patterns[col].append((row_num, cell_value))
    
    # Print summary
    print(f"\nColumns that appear to contain quantities:")
    for col, examples in quantity_patterns.items():
        col_header = sheet.cell(row=1, column=col).value or f"Column {col}"
        print(f"\n  Column {col} ('{col_header}'):")
        print(f"    Found {len(examples)} quantity values")
        print(f"    First 5 examples: {examples[:5]}")
    
    print(f"\n" + "="*80)
    print(f"SUMMARY:")
    print(f"="*80)
    print(f"Total rows with potential quantities found: {len(all_rows_with_quantities)}")
    print(f"Row numbers with quantities: {sorted(all_rows_with_quantities[:20])}...")
    
    # Look for specific patterns that might indicate quantities
    print(f"\n" + "="*80)
    print(f"SEARCHING FOR QUANTITY PATTERNS:")
    print(f"="*80)
    
    # Check for common quantity column names
    quantity_keywords = ['quantity', 'qty', 'amount', 'count', 'units', 'pieces', 'pcs']
    
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and any(keyword in str(header).lower() for keyword in quantity_keywords):
            print(f"\nPotential quantity column found: Column {col} - '{header}'")
            
            # Count non-zero values in this column
            non_zero_count = 0
            examples = []
            for row in range(2, sheet.max_row + 1):
                value = sheet.cell(row=row, column=col).value
                if value and value != 0:
                    non_zero_count += 1
                    if len(examples) < 5:
                        examples.append((row, value))
            
            print(f"  Non-zero values in this column: {non_zero_count}")
            print(f"  Examples: {examples}")
    
    workbook.close()

if __name__ == "__main__":
    file_path = r"C:\Users\abaza\Downloads\TESTFILE.xlsx"
    analyze_excel_file(file_path)